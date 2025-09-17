


import openpyxl
import os
import logging
import re

from src.excel_mapper import get_excel_path

logger = logging.getLogger(__name__)


def normalise_text(text: str) -> str:
    """
        Normalises text for comparison:
        - Strips leading/trailing spaces
        - Converts to lowercase
        - Replaces multiple spaces with single space
        - Removes some punctuation
        """
    text = text.strip().lower()
    text = re.sub(r'\s+', ' ', text)  # collapse multiple spaces    
    text = re.sub(r'[^\w\s]', '', text)  # remove punctuation
    return text


def update_excel_with_submission(parsed_email: dict):
    """
    Update Excel with parsed submission details.
    Checks if student and assignment already exist, otherwise creates them.
    """
    course_name = parsed_email.get("course_name")
    assignment = parsed_email.get("assignment")
    student_name = parsed_email.get("user")

    excel_path = get_excel_path(course_name)

    if not excel_path:
        logger.error(f"No Excel mapping found for course: {course_name}")
        return False
    
    # Insure file exists
    if not os.path.exists(excel_path):
        logger.error(f"Excel file does not exist: {excel_path}")
        return False
    
    # load workbook 
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active  # assuming we want the active sheet

    assignment_col = None
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col).value).strip().lower()
        if normalise_text(header) == normalise_text(assignment):
            assignment_col = col
            break
    
    if not assignment_col:
        # Create new assignment column at the end
        assignment_col = ws.max_column + 1
        ws.cell(row=1, column=assignment_col, value=assignment)
        logger.info(f"Added new assignment column: {assignment} at column {assignment_col}")

    # Find or create studnet row
    student_row = None
    for row in range(2, ws.max_row + 1):
        full_name = str(ws.cell(row=row, column=1).value).strip().lower()
        if normalise_text(full_name) == normalise_text(student_name):
            student_row = row
            break

    if not student_row:
        # Create new student row at the end
        student_row = ws.max_row + 1
        ws.cell(row=student_row, column=1, value=student_name)
        ws.cell(row=student_row, column=2, value=course_name)
        logger.info(f"Added new student: {student_name} at row {student_row}")

    # Mark assignment as submitted
    ws.cell(row=student_row, column=assignment_col, value="Done")

    wb.save(excel_path)
    logger.info(f"Updated Excel file: {excel_path} for student: {student_name}, assignment: {assignment}")

    return True



    # #find the next empty row
    # next_row = ws.max_row + 1

    # # write parsed data (simple version)
    # ws.cell(row=next_row, column=1, value=parsed_email.get("submission_id"))
    # ws.cell(row=next_row, column=2, value=parsed_email.get("course_code"))
    # ws.cell(row=next_row, column=3, value=parsed_email.get("course_name"))
    # ws.cell(row=next_row, column=4, value=parsed_email.get("assignment"))
    # ws.cell(row=next_row, column=5, value=parsed_email.get("user")) 

    # # Save the workbook
    # wb.save(excel_path)
    # logger.info(f"Updated Excel file: {excel_path} with submission ID: {parsed_email.get('submission_id')}")
    
    # return True

